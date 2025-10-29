"""
Servicios auxiliares para el sistema de nutrición
"""
import mercadopago
from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import io
import logging

logger = logging.getLogger(__name__)


class MercadoPagoService:
    """Servicio para integración con MercadoPago"""
    
    def __init__(self):
        # Obtener access token desde settings (configurar en .env)
        access_token = getattr(settings, 'MERCADOPAGO_ACCESS_TOKEN', None)
        self.sdk = mercadopago.SDK(access_token) if access_token else None
    
    def create_payment_preference(self, payment_data):
        """
        Crear una preferencia de pago en MercadoPago
        
        payment_data debe contener:
        - title: Título del pago
        - quantity: Cantidad
        - unit_price: Precio unitario
        - payer_email: Email del pagador
        - external_reference: Referencia externa (ID del pago en DB)
        """
        if not self.sdk:
            logger.error("MercadoPago SDK no configurado")
            return {'success': False, 'error': 'SDK no configurado'}
        
        try:
            preference_data = {
                "items": [
                    {
                        "title": payment_data.get('title', 'Consulta Nutricional'),
                        "quantity": payment_data.get('quantity', 1),
                        "unit_price": float(payment_data.get('unit_price', 0))
                    }
                ],
                "payer": {
                    "email": payment_data.get('payer_email')
                },
                "external_reference": str(payment_data.get('external_reference', '')),
                "back_urls": {
                    "success": payment_data.get('success_url', 'http://localhost:5175/payment/success'),
                    "failure": payment_data.get('failure_url', 'http://localhost:5175/payment/failure'),
                    "pending": payment_data.get('pending_url', 'http://localhost:5175/payment/pending')
                },
                "auto_return": "approved",
                "notification_url": payment_data.get('notification_url', 'http://localhost:8000/api/payments/webhook/'),
            }
            
            preference_response = self.sdk.preference().create(preference_data)
            preference = preference_response["response"]
            
            return {
                'success': True,
                'preference_id': preference['id'],
                'init_point': preference['init_point'],  # URL para pagar
                'sandbox_init_point': preference.get('sandbox_init_point')  # URL para sandbox
            }
            
        except Exception as e:
            logger.error(f"Error al crear preferencia de MercadoPago: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def get_payment_info(self, payment_id):
        """Obtener información de un pago"""
        if not self.sdk:
            return {'success': False, 'error': 'SDK no configurado'}
        
        try:
            payment_response = self.sdk.payment().get(payment_id)
            payment = payment_response["response"]
            
            return {
                'success': True,
                'payment': payment
            }
        except Exception as e:
            logger.error(f"Error al obtener info de pago: {str(e)}")
            return {'success': False, 'error': str(e)}


class EmailNotificationService:
    """Servicio para envío de notificaciones por email"""
    
    @staticmethod
    def send_appointment_reminder(appointment):
        """Enviar recordatorio de cita"""
        patient = appointment.patient
        patient_user = patient.person.user
        nutritionist = appointment.nutritionist
        
        subject = f'Recordatorio: Cita con {nutritionist.get_full_name()}'
        
        # Contexto para el template
        context = {
            'patient_name': patient_user.get_full_name(),
            'nutritionist_name': nutritionist.get_full_name(),
            'appointment_date': appointment.appointment_date.strftime('%d/%m/%Y'),
            'appointment_time': appointment.appointment_time.strftime('%H:%M'),
            'consultation_type': appointment.get_consultation_type_display(),
            'notes': appointment.notes
        }
        
        # Mensaje en texto plano
        message = f"""
Hola {context['patient_name']},

Te recordamos que tienes una cita programada:

Nutricionista: {context['nutritionist_name']}
Tipo de consulta: {context['consultation_type']}
Fecha: {context['appointment_date']}
Hora: {context['appointment_time']}

{f'Notas: {context["notes"]}' if context['notes'] else ''}

Por favor, confirma tu asistencia.

Saludos,
Sistema de Nutrición
        """
        
        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@nutricion.com',
                recipient_list=[patient_user.email],
                fail_silently=False,
            )
            logger.info(f"Recordatorio enviado a {patient_user.email} para cita #{appointment.id}")
            return True
        except Exception as e:
            logger.error(f"Error al enviar recordatorio: {str(e)}")
            return False
    
    @staticmethod
    def send_payment_confirmation(payment):
        """Enviar confirmación de pago"""
        patient = payment.patient
        patient_user = patient.person.user
        nutritionist = payment.nutritionist
        
        subject = 'Confirmación de Pago Recibido'
        
        message = f"""
Hola {patient_user.get_full_name()},

Confirmamos que hemos recibido tu pago:

Monto: ${payment.amount}
Método: {payment.get_payment_method_display()}
Descripción: {payment.description}
Fecha: {payment.payment_date.strftime('%d/%m/%Y %H:%M') if payment.payment_date else 'Pendiente'}

Gracias por tu pago.

Saludos,
{nutritionist.get_full_name()}
Sistema de Nutrición
        """
        
        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@nutricion.com',
                recipient_list=[patient_user.email],
                fail_silently=False,
            )
            logger.info(f"Confirmación de pago enviada a {patient_user.email}")
            return True
        except Exception as e:
            logger.error(f"Error al enviar confirmación de pago: {str(e)}")
            return False
    
    @staticmethod
    def send_plan_renewal_reminder(patient, nutritionist):
        """Enviar recordatorio de renovación de plan"""
        patient_user = patient.person.user
        
        subject = 'Recordatorio: Renovación de Plan Nutricional'
        
        message = f"""
Hola {patient_user.get_full_name()},

Tu plan nutricional está próximo a vencer. 

Te recomendamos agendar una cita de seguimiento con {nutritionist.get_full_name()} para renovar tu plan y evaluar tu progreso.

Saludos,
Sistema de Nutrición
        """
        
        try:
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL if hasattr(settings, 'DEFAULT_FROM_EMAIL') else 'noreply@nutricion.com',
                recipient_list=[patient_user.email],
                fail_silently=False,
            )
            logger.info(f"Recordatorio de renovación enviado a {patient_user.email}")
            return True
        except Exception as e:
            logger.error(f"Error al enviar recordatorio de renovación: {str(e)}")
            return False


class ReportService:
    """Servicio para generación de reportes en PDF y Excel"""
    
    @staticmethod
    def generate_patient_report_pdf(patient):
        """Generar reporte completo de paciente en PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a5490'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        title = Paragraph(f"Reporte de Paciente: {patient.person.user.get_full_name()}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Información personal
        data_personal = [
            ['Información Personal', ''],
            ['DNI:', patient.person.user.dni],
            ['Email:', patient.person.user.email],
            ['Teléfono:', patient.person.phone],
            ['Fecha de Nacimiento:', patient.person.birth_date.strftime('%d/%m/%Y') if patient.person.birth_date else 'N/A'],
            ['Género:', patient.person.get_gender_display() if patient.person.gender else 'N/A'],
        ]
        
        table_personal = Table(data_personal, colWidths=[2*inch, 4*inch])
        table_personal.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5490')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table_personal)
        elements.append(Spacer(1, 0.3*inch))
        
        # Historial de consultas
        consultations = patient.consultations.all().order_by('-date')[:10]
        
        if consultations:
            subtitle = Paragraph("Últimas Consultas", styles['Heading2'])
            elements.append(subtitle)
            elements.append(Spacer(1, 0.1*inch))
            
            data_consultations = [['Fecha', 'Tipo', 'Peso (kg)', 'IMC']]
            
            for consultation in consultations:
                measurements = getattr(consultation, 'measurements', None)
                data_consultations.append([
                    consultation.date.strftime('%d/%m/%Y'),
                    consultation.get_consultation_type_display(),
                    f"{measurements.weight:.2f}" if measurements else 'N/A',
                    f"{measurements.bmi:.2f}" if measurements and measurements.bmi else 'N/A'
                ])
            
            table_consultations = Table(data_consultations, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1.5*inch])
            table_consultations.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5490')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table_consultations)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_patient_evolution_excel(patient):
        """Generar reporte de evolución del paciente en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Evolución Paciente"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a5490", end_color="1a5490", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Información del paciente
        ws['A1'] = 'REPORTE DE EVOLUCIÓN - PACIENTE'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        ws['A3'] = 'Nombre:'
        ws['B3'] = patient.person.user.get_full_name()
        ws['A4'] = 'DNI:'
        ws['B4'] = patient.person.user.dni
        ws['A5'] = 'Email:'
        ws['B5'] = patient.person.user.email
        
        # Headers de evolución
        row = 7
        headers = ['Fecha', 'Tipo Consulta', 'Peso (kg)', 'Altura (m)', 'IMC', 'ICC', 'TMB', 'GET']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos de consultas
        consultations = patient.consultations.all().order_by('date')
        row = 8
        
        for consultation in consultations:
            measurements = getattr(consultation, 'measurements', None)
            person = patient.person
            
            # Calcular edad si hay fecha de nacimiento
            age = None
            tmb = None
            get_value = None
            
            if person.birth_date and person.gender and measurements:
                from datetime import date
                age = (date.today() - person.birth_date).days // 365
                tmb = measurements.calculate_tmb(age, person.gender)
                get_value = measurements.calculate_get(age, person.gender)
            
            ws.cell(row=row, column=1, value=consultation.date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=consultation.get_consultation_type_display())
            ws.cell(row=row, column=3, value=measurements.weight if measurements else 'N/A')
            ws.cell(row=row, column=4, value=measurements.height if measurements else 'N/A')
            ws.cell(row=row, column=5, value=measurements.bmi if measurements and measurements.bmi else 'N/A')
            ws.cell(row=row, column=6, value=measurements.waist_hip_ratio if measurements and measurements.waist_hip_ratio else 'N/A')
            ws.cell(row=row, column=7, value=tmb if tmb else 'N/A')
            ws.cell(row=row, column=8, value=get_value if get_value else 'N/A')
            
            row += 1
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_monthly_report_pdf(nutritionist, start_date, end_date):
        """Generar reporte mensual de actividad del nutricionista"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph(f"Reporte Mensual - {nutritionist.get_full_name()}", styles['Heading1'])
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        period = Paragraph(f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}", styles['Normal'])
        elements.append(period)
        elements.append(Spacer(1, 0.3*inch))
        
        # Estadísticas
        from .models import Appointment, Payment, Consultation
        from django.db import models as db_models
        
        appointments_count = Appointment.objects.filter(
            nutritionist=nutritionist,
            appointment_date__range=[start_date, end_date]
        ).count()
        
        consultations_count = Consultation.objects.filter(
            nutritionist=nutritionist,
            date__range=[start_date, end_date]
        ).count()
        
        payments_sum = Payment.objects.filter(
            nutritionist=nutritionist,
            payment_date__range=[start_date, end_date],
            status='approved'
        ).aggregate(total=db_models.Sum('amount'))['total'] or 0
        
        data_stats = [
            ['Estadísticas del Período', ''],
            ['Total de Citas:', str(appointments_count)],
            ['Total de Consultas:', str(consultations_count)],
            ['Ingresos Totales:', f'${payments_sum:.2f}'],
        ]
        
        table_stats = Table(data_stats, colWidths=[3*inch, 2*inch])
        table_stats.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5490')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table_stats)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

